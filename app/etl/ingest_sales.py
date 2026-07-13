"""Streaming ETL for IP + OP Sales (billing) → small aggregate parquet tables.

The raw sales files are large (IP xlsx up to ~280MB) and have inconsistent schemas
month to month (29 vs 41 cols, different sheet names, .xlsx and .xlsb). We stream
each file row-by-row (never hold a whole file in memory), auto-detect the data
sheet + header, harmonise columns, keep the Dec-2025..May-2026 window, and
accumulate aggregates the dashboards need:

  sales_monthly.parquet        month × patient(IP/OP) → revenue, cost, qty, lines
  sales_by_hospital.parquet    hospital(code) → revenue, cost, qty
  sales_by_manufacturer.parquet manufacturer → revenue, cost, qty
  sales_by_material.parquet    material → desc, group, revenue, cost, qty
  sales_totals.parquet         patient → revenue, cost, qty, lines

Revenue = TOTALMRP (billed), Cost = TOTALCOSTPRICE, Margin = revenue − cost (real).
"""
from __future__ import annotations
import os, glob
from collections import defaultdict
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook

try:
    from pyxlsb import open_workbook as open_xlsb
except Exception:
    open_xlsb = None

HERE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # backend root
RAW = os.environ.get("SALES_RAW", "/Users/shivanshdarshan/Documents/Profession/bidezy/codebase/Bidezy 2")
KPI = os.path.join(HERE, "app", "data", "kpi")
LIMIT = int(os.environ.get("SALES_LIMIT", "0"))  # >0 = only first N data rows/file (prototype)

KEEP = {(2025, 12), (2026, 1), (2026, 2), (2026, 3), (2026, 4), (2026, 5)}  # match the rest of the data
EXCEL_BASE = datetime(1899, 12, 30)

FIELDS = {
    "date": ["SALESDATE", "BILLDATE"],
    "mat": ["MATERIALCODE", "ITEMCODE"],
    "desc": ["MATERIALLINEITEM", "ITEMNAME"],
    "qty": ["QTY"],
    "rev": ["TOTALMRP"],
    "cost": ["TOTALCOSTPRICE"],
    "mfr": ["MANUFACTURERNAME"],
    "inv": ["SALESINVOICENO", "BILLNO"],
    "loc": ["LOCATIONNAME", "STORENAME"],
}
REQUIRED = ("date", "qty", "rev", "cost", "mat")


def parse_ym(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return (v.year, v.month)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            d = EXCEL_BASE + timedelta(days=float(v)); return (d.year, d.month)
        except Exception:
            return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            d = datetime.strptime(s[:19] if len(s) >= 19 and " " in s else s.split(" ")[0], fmt)
            return (d.year, d.month)
        except Exception:
            pass
    try:
        d = EXCEL_BASE + timedelta(days=float(s)); return (d.year, d.month)
    except Exception:
        return None


def col_map(header):
    H = [str(h).strip().upper() if h is not None else "" for h in header]
    idx = {}
    for f, cands in FIELDS.items():
        for c in cands:
            if c in H:
                idx[f] = H.index(c); break
    return idx if all(k in idx for k in REQUIRED) else None


def _num(v):
    try:
        return float(v)
    except Exception:
        return None


def iter_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sh in wb.sheetnames:
            ws = wb[sh]
            it = ws.iter_rows(values_only=True)
            idx = None
            for _ in range(8):
                try:
                    row = next(it)
                except StopIteration:
                    break
                idx = col_map(row)
                if idx:
                    break
            if not idx:
                continue
            for row in it:
                yield idx, row
            return
    finally:
        wb.close()


def iter_xlsb(path):
    if open_xlsb is None:
        return
    with open_xlsb(path) as wb:
        for sh in wb.sheets:
            with wb.get_sheet(sh) as sheet:
                idx = None
                for r in sheet.rows():
                    vals = [c.v for c in r]
                    if idx is None:
                        idx = col_map(vals)
                        continue
                    yield idx, vals
                if idx is not None:
                    return


def records(path):
    it = iter_xlsb(path) if path.lower().endswith(".xlsb") else iter_xlsx(path)
    n = 0
    for idx, row in it:
        def g(f):
            i = idx.get(f)
            return row[i] if (i is not None and i < len(row)) else None
        yield g
        n += 1
        if LIMIT and n >= LIMIT:
            return


def hosp_of(inv, loc):
    if inv:
        s = str(inv)
        if "-" in s:
            return s.split("-")[0].strip().upper()
    return str(loc).strip() if loc else "UNKNOWN"


def run():
    tot = defaultdict(lambda: defaultdict(float))
    month = defaultdict(lambda: defaultdict(float))
    hosp = defaultdict(lambda: defaultdict(float))
    mfr = defaultdict(lambda: defaultdict(float))
    mat = defaultdict(lambda: defaultdict(float))
    matdesc = {}

    def files(sub):
        d = os.path.join(RAW, sub)
        return sorted(glob.glob(d + "/*.xlsx") + glob.glob(d + "/*.xlsb"))

    sources = [("IP", f) for f in files("IP Sales")] + [("OP", f) for f in files("OP Sales")]
    for patient, path in sources:
        kept = 0; seen = 0
        for g in records(path):
            seen += 1
            ym = parse_ym(g("date"))
            if ym not in KEEP:
                continue
            rev = _num(g("rev")); cost = _num(g("cost")); qty = _num(g("qty"))
            if rev is None:
                continue
            cost = cost or 0.0; qty = qty or 0.0
            code = str(g("mat")).strip() if g("mat") is not None else ""
            if not code or code.lower() == "none":
                continue
            ymk = f"{ym[0]:04d}-{ym[1]:02d}"
            for bucket, key in ((tot, patient), (month, (patient, ymk)),
                                (hosp, hosp_of(g("inv"), g("loc"))), (mat, code)):
                bucket[key]["revenue"] += rev; bucket[key]["cost"] += cost
                bucket[key]["qty"] += qty; bucket[key]["lines"] += 1
            m = g("mfr")
            if m and str(m).strip().lower() not in ("none", "nan", ""):
                mm = mfr[str(m).strip().title()]
                mm["revenue"] += rev; mm["cost"] += cost; mm["qty"] += qty; mm["lines"] += 1
            matdesc.setdefault(code, str(g("desc")) if g("desc") else code)
            kept += 1
        print(f"[sales] {patient:2s} {os.path.basename(path)[:38]:38s} seen={seen:>8,} kept={kept:>8,}", flush=True)

    # category from material master
    grp = {}
    try:
        dm = pd.read_parquet(os.path.join(KPI, "..", "curated", "dim_material.parquet"))
    except Exception:
        dm = None
    if dm is None:
        try:
            dm = pd.read_parquet(os.path.join(HERE, "app", "data", "curated", "dim_material.parquet"))
        except Exception:
            dm = None
    if dm is not None and "material" in dm and "material_group" in dm:
        grp = dict(zip(dm["material"].astype(str), dm["material_group"].astype(str)))

    os.makedirs(KPI, exist_ok=True)
    pd.DataFrame([{"patient": k, **v} for k, v in tot.items()]).to_parquet(os.path.join(KPI, "sales_totals.parquet"), index=False)
    pd.DataFrame([{"patient": k[0], "month": k[1], **v} for k, v in month.items()]).to_parquet(os.path.join(KPI, "sales_monthly.parquet"), index=False)
    pd.DataFrame([{"hospital": k, **v} for k, v in hosp.items()]).to_parquet(os.path.join(KPI, "sales_by_hospital.parquet"), index=False)
    pd.DataFrame([{"manufacturer": k, **v} for k, v in mfr.items()]).to_parquet(os.path.join(KPI, "sales_by_manufacturer.parquet"), index=False)
    pd.DataFrame([{"material": k, "desc": matdesc.get(k, k), "group": grp.get(k, ""), **v} for k, v in mat.items()]).to_parquet(os.path.join(KPI, "sales_by_material.parquet"), index=False)

    grand = sum(v["revenue"] for v in tot.values())
    gcost = sum(v["cost"] for v in tot.values())
    print(f"\n[sales] DONE. revenue=₹{grand/1e7:.1f}Cr cost=₹{gcost/1e7:.1f}Cr margin=₹{(grand-gcost)/1e7:.1f}Cr ({(grand-gcost)/grand*100:.1f}%)")
    print(f"[sales] patients: " + ", ".join(f"{k}=₹{v['revenue']/1e7:.1f}Cr" for k, v in tot.items()))
    print(f"[sales] hospitals={len(hosp)} manufacturers={len(mfr)} materials={len(mat)}")


if __name__ == "__main__":
    run()
